# python 2.7
import openpyxl

#############
# constants
#############
# the column constants need to be changed if format changes
COLUMN_ZIPCODE = 16
COLUMN_GRADE = 17
COLUMN_SCHOOL = 18

#############
# functions
#############
def update_grade(grade):
    if not grade or grade == '':
        new_grade = ''
    elif grade == 'PS':
        new_grade = 'PreK'
    elif grade == 'PreK':
        new_grade = 'K'
    elif grade == 'K':
        new_grade = 1
    elif grade == 12:
        # remove row
        new_grade = 'REMOVE'
    else:
        try:
            int_grade = int(grade)
            new_grade = int_grade+1
        except ValueError:
            print('Unknown value for grade: "%s"' % grade)
#    print("old grade: %s"%grade)
#    print("new grade: %s"%new_grade)
    return new_grade

def fix_zipcode(zipcode):
    if not zipcode or zipcode == '':
        new_zipcode = ''
    else:
        if len(str(zipcode)) == 4:
            new_zipcode = '0'+str(zipcode)
        else:
            new_zipcode = str(zipcode)
            print("zip code not modified: %s"%zipcode)
#    print("old zip: %s"%zipcode)
#    print("new zip: %s"%new_zipcode)
    return new_zipcode

def fix_school(school, grade):
    new_school = school
    if grade in [6,7,8] and school != 'HMS':
        if school in ['Haddon', 'Central', 'Tatem']:
            new_school = 'HMS'
#            print("Updated school from %s to %s (%d-th)" %(school, new_school, grade))
        else:
            print("unusual school %s grade %d" %(school,grade))
    if grade in [9,10,11,12] and school != 'HMHS':
        if school == 'HMS':
            new_school = 'HMHS'
#            print("Updated school from %s to %s (%d-th)" %(school, new_school, grade))
        else:
            print("unusual school %s grade %d" %(school,grade))
    return new_school


#############
# main
#############
book = openpyxl.load_workbook('directory.xlsx')
sheet = book.active

# data starts on 2nd row
for i in range(2,sheet.max_row):
    # zip code
    cell_zipcode = sheet.cell(row=i, column=COLUMN_ZIPCODE)
    cell_zipcode.value = fix_zipcode(cell_zipcode.value)
    # grade
    cell_grade = sheet.cell(row=i, column=COLUMN_GRADE)
    new_grade = update_grade(cell_grade.value)
    cell_grade.value = new_grade
    # school
    cell_school = sheet.cell(row=i, column=COLUMN_SCHOOL)
    cell_school.value = fix_school(cell_school.value, new_grade)
book.save('directory-new.xlsx')
